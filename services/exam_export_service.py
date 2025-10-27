import pandas as pd
from datetime import datetime
from typing import Optional, List
from services.db import fetch_all
import os

class ExamExportService:
    """Sınav programlarını Excel'e aktarma servisi"""
    
    def __init__(self):
        pass
    
    def export_exam_schedule_to_excel(self, schedule_id: int, output_path: Optional[str] = None) -> str:
        # 1. Program bilgilerini al
        schedule = fetch_all("""
            SELECT es.*, d.name as department_name
            FROM exam_schedules es
            JOIN departments d ON es.department_id = d.id
            WHERE es.id = %s
        """, [schedule_id])
        
        if not schedule:
            raise ValueError(f"Schedule ID {schedule_id} bulunamadi!")
        
        schedule = schedule[0]
        
        # 2. Sınavları al
        exams = fetch_all("""
            SELECT 
                e.id,
                e.exam_date,
                e.start_time,
                e.end_time,
                e.duration,
                e.student_count,
                c.code as course_code,
                c.name as course_name,
                c.instructor,
                c.grade,
                c.is_elective,
                cl.code as classroom_code,
                cl.name as classroom_name,
                cl.capacity as classroom_capacity
            FROM exams e
            JOIN courses c ON e.course_id = c.id
            LEFT JOIN exam_classrooms ec ON e.id = ec.exam_id
            LEFT JOIN classrooms cl ON ec.classroom_id = cl.id
            WHERE e.schedule_id = %s
            ORDER BY e.exam_date, e.start_time, c.code
        """, [schedule_id])
        
        if not exams:
            raise ValueError("Bu programa ait sinav bulunamadi!")
        
        
        # 3. Excel için veri hazırla
        excel_data = []
        previous_date = None
        previous_time = None
        
        for exam in exams:
            # Tarih formatla
            exam_date = exam['exam_date'].strftime('%d.%m.%Y') if hasattr(exam['exam_date'], 'strftime') else str(exam['exam_date'])
            exam_day = self._get_day_name(exam['exam_date'])
            
            # Saat formatla
            start_time = exam['start_time'].strftime('%H:%M') if hasattr(exam['start_time'], 'strftime') else str(exam['start_time'])
            end_time = exam['end_time'].strftime('%H:%M') if hasattr(exam['end_time'], 'strftime') else str(exam['end_time'])
            time_slot = f"{start_time} - {end_time}"
            
            # Aynı tarihse boş bırak (birleştirme için)
            if exam_date == previous_date:
                display_date = ''
                display_day = ''
            else:
                display_date = exam_date
                display_day = exam_day
                previous_date = exam_date
                previous_time = None  # Yeni gün, saat kontrolünü sıfırla
            
            # Aynı saat aralığındaysa boş bırak (birleştirme için)
            if time_slot == previous_time:
                display_time = ''
            else:
                display_time = time_slot
                previous_time = time_slot
            
            excel_data.append({
                'Tarih': display_date,
                'Gün': display_day,
                'Saat': display_time,
                'Ders Kodu': exam['course_code'],
                'Ders Adı': exam['course_name'],
                'Sınıf': f"{exam['grade']}. Sınıf",  # YENİ SÜTUN
                'Öğretim Üyesi': exam['instructor'] or '',
                'Derslik': exam['classroom_name'] or 'Atanmadı'
            })
        
        # 4. DataFrame oluştur
        df = pd.DataFrame(excel_data)
        
        # 5. Dosya yolu belirle
        if not output_path:
            # Otomatik dosya adı oluştur
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            safe_name = schedule['name'].replace(' ', '_').replace('/', '-')
            filename = f"{safe_name}_{schedule['exam_type']}_{timestamp}.xlsx"
            output_path = os.path.join(os.getcwd(), filename)
        
        # 6. Excel'e yaz
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Ana sayfa - Program detayları (2. satırdan başla, 1. satır başlık için)
            df.to_excel(writer, sheet_name='Sinav Programi', index=False, startrow=2)
            
            # Bilgi sayfası
            info_data = {
                'Bilgi': [
                    'Program Adı',
                    'Sınav Türü',
                    'Bölüm',
                    'Başlangıç Tarihi',
                    'Bitiş Tarihi',
                    'Toplam Sınav',
                    'Oluşturulma Tarihi'
                ],
                'Değer': [
                    schedule['name'],
                    schedule['exam_type'],
                    schedule['department_name'],
                    str(schedule['start_date']),
                    str(schedule['end_date']),
                    len(exams),
                    schedule['created_at'].strftime('%d.%m.%Y %H:%M') if hasattr(schedule['created_at'], 'strftime') else str(schedule['created_at'])
                ]
            }
            
            df_info = pd.DataFrame(info_data)
            df_info.to_excel(writer, sheet_name='Program Bilgileri', index=False)
        
        # 7. Excel formatting (opsiyonel - güzel görünüm için)
        self._format_excel(output_path, schedule['department_name'], schedule['exam_type'])
        
        return output_path
    
    def _get_day_name(self, date_obj) -> str:
        """Tarihten gün adını al"""
        if hasattr(date_obj, 'weekday'):
            days = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
            return days[date_obj.weekday()]
        return ''
    
    def _format_excel(self, file_path: str, department_name: str = "", exam_type: str = ""):
        """Excel dosyasını formatla (güzel görünüm için)"""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = load_workbook(file_path)

            if 'Sinav Programi' in wb.sheetnames:
                ws = wb['Sinav Programi']

                ws.merge_cells('A1:H1')
                title_cell = ws['A1']
                title_text = f"{department_name.upper()} - {exam_type.upper()} SINAV PROGRAMI"
                title_cell.value = title_text
                title_cell.font = Font(size=18, bold=True, italic=True, color='FFFFFF')
                title_cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[1].height = 30

                ws.row_dimensions[2].height = 10

                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_font = Font(bold=True, italic=True, color='FFFFFF', size=12)
                
                for cell in ws[3]:
                    if cell.value:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                ws.row_dimensions[3].height = 25

                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=8):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                self._merge_same_dates(ws, start_row=4, date_col=1)

                self._merge_same_times(ws, start_row=4, time_col=3)

                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 18
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 40
                ws.column_dimensions['F'].width = 12
                ws.column_dimensions['G'].width = 30
                ws.column_dimensions['H'].width = 15
            
            wb.save(file_path)
            pass
            
        except Exception as e:
            pass
    
    def _merge_same_dates(self, ws, start_row: int, date_col: int):
        """Aynı tarihleri birleştir"""
        try:
            from openpyxl.styles import Alignment
            
            current_date = None
            merge_start = None
            
            for row_idx in range(start_row, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=date_col).value
                
                if cell_value and cell_value.strip():
                    if merge_start and merge_start < row_idx - 1:
                        ws.merge_cells(f'A{merge_start}:A{row_idx-1}')
                        ws.merge_cells(f'B{merge_start}:B{row_idx-1}')
                        ws.cell(row=merge_start, column=1).alignment = Alignment(horizontal='center', vertical='center')
                        ws.cell(row=merge_start, column=2).alignment = Alignment(horizontal='center', vertical='center')

                    current_date = cell_value
                    merge_start = row_idx
                elif current_date:
                    pass

            if merge_start and merge_start < ws.max_row:
                ws.merge_cells(f'A{merge_start}:A{ws.max_row}')
                ws.merge_cells(f'B{merge_start}:B{ws.max_row}')
                ws.cell(row=merge_start, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=merge_start, column=2).alignment = Alignment(horizontal='center', vertical='center')
                
        except Exception as e:
            pass
    
    def _merge_same_times(self, ws, start_row: int, time_col: int):
        """Aynı saat aralıklarını birleştir"""
        try:
            from openpyxl.styles import Alignment
            
            current_time = None
            merge_start = None
            
            for row_idx in range(start_row, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=time_col).value
                
                if cell_value and str(cell_value).strip():
                    if merge_start and merge_start < row_idx - 1:
                        ws.merge_cells(start_row=merge_start, start_column=time_col, 
                                      end_row=row_idx-1, end_column=time_col)

                        ws.cell(row=merge_start, column=time_col).alignment = Alignment(
                            horizontal='center', vertical='center'
                        )

                    current_time = cell_value
                    merge_start = row_idx
                elif current_time:
                    pass

            if merge_start and merge_start < ws.max_row:
                ws.merge_cells(start_row=merge_start, start_column=time_col,
                              end_row=ws.max_row, end_column=time_col)
                ws.cell(row=merge_start, column=time_col).alignment = Alignment(
                    horizontal='center', vertical='center'
                )
                
        except Exception as e:
            pass
    
    def export_all_schedules(self, department_id: int, output_dir: Optional[str] = None) -> List[str]:

        schedules = fetch_all("""
            SELECT id, name FROM exam_schedules 
            WHERE department_id = %s 
            ORDER BY created_at DESC
        """, [department_id])
        
        exported_files = []
        
        for schedule in schedules:
            try:
                file_path = self.export_exam_schedule_to_excel(
                    schedule['id'],
                    output_path=os.path.join(output_dir, f"{schedule['name']}.xlsx") if output_dir else None
                )
                exported_files.append(file_path)
            except Exception as e:
                pass
        
        return exported_files


exam_export_service = ExamExportService()

